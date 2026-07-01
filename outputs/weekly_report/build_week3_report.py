from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


output_dir = Path(r"C:\Users\26474\Documents\Unreal Projects\TheGameOf2D\outputs\weekly_report")
output_dir.mkdir(parents=True, exist_ok=True)
output = Path(r"C:\Users\26474\OneDrive\文档\1组-第3周总结报告.xlsx")
workspace_output = output_dir / "1组-第3周总结报告.xlsx"

wb = Workbook()
overview = wb.active
overview.title = "周报总览"
modules = wb.create_sheet("模块状态")
summary = wb.create_sheet("本周小结")
personal = wb.create_sheet("个人小结")
meeting = wb.create_sheet("讨论纪要与截图")

thin = Side(style="thin", color="D9E2EC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
title_fill = PatternFill("solid", fgColor="1F4E5F")
section_fill = PatternFill("solid", fgColor="EAF3F5")
progress_fill = PatternFill("solid", fgColor="FFF3CD")
done_fill = PatternFill("solid", fgColor="D9EAD3")
white_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
header_font = Font(name="Microsoft YaHei", color="1F2937", bold=True, size=11)
body_font = Font(name="Microsoft YaHei", color="1F2937", size=10)


def merge(ws, range_name, value, fill=None, font=None, align=None):
    ws.merge_cells(range_name)
    cell = ws[range_name.split(":")[0]]
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = align or Alignment(vertical="top", wrap_text=True)
    for row in ws[range_name]:
        for c in row:
            c.border = border


def style_range(ws, cell_range, fill=None, font=None, align=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = copy(font)
            cell.border = border
            cell.alignment = align or Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


for ws in wb.worksheets:
    ws.freeze_panes = "A1"

# 周报总览
set_widths(overview, [18, 34, 34, 34])
merge(overview, "A1:D1", "每周总结报告", title_fill, white_font, Alignment(horizontal="center", vertical="center"))
overview.row_dimensions[1].height = 30
merge(overview, "A2:D2", "《墨灯守巷》第三周项目进度总结", None, Font(name="Microsoft YaHei", bold=True, size=14), Alignment(horizontal="center", vertical="center"))
overview["A4"], overview["B4"] = "组别", "第1组"
overview["A5"], overview["B5"] = "游戏名", "墨灯守巷"
overview["A6"], overview["B6"] = "时间周期", "第三周"
style_range(overview, "A4:A6", section_fill, header_font)
style_range(overview, "B4:D6")
merge(overview, "A8:D8", "本周进度概览", title_fill, Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=12), Alignment(vertical="center"))
merge(
    overview,
    "A9:D13",
    "本周开发重点围绕项目计划书中第三周阶段验收标准展开，即完善敌人 AI、波次刷怪、投射物或击退、玩家血量和基础 UI，并接入第一批音效，推动第一关从第二周最小玩法闭环进入 Alpha 雏形阶段。结合当前 GitHub 最新进度，本周主线分工为：邢增浩继续负责主角更新，重点推进 Tomoe 主角移动、跳跃下落、武士刀攻击窗口和地面连招；黄帝尧继续负责怪物更新，重点推进敌人配置、敌人视觉资源、敌人生成与战斗表现，并同步修复灯笼交互和可见性问题。当前项目已经在主角战斗表现、怪物表现和第一关玩法支撑上明显向第三周目标推进。",
)
merge(overview, "A15:D15", "当前风险与后续重点", title_fill, Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=12), Alignment(vertical="center"))
merge(
    overview,
    "A16:D20",
    "总体来看，第三周工作正在从“能玩通”的最小闭环转向“可展示”的 Alpha 版本。当前仍需要重点补齐玩家血量、灯火弹或击退效果、基础 UMG UI、音效接入和一次组内测试记录；主角 Tomoe 动画与连招需要继续在蓝图和 C++ 动画通知之间联调，怪物侧需要继续确认普通敌人、快速敌人、自爆敌人的行为区分和波次节奏。下周应优先保证第一关 Alpha 能稳定演示，再进入第二关、小 Boss、关卡切换和过场准备。",
)
style_range(overview, "A1:D20")
for r in range(9, 21):
    overview.row_dimensions[r].height = 31

# 模块状态
set_widths(modules, [16, 16, 26, 84])
modules.append(["模块状态", "负责人", "模块", "详细说明"])
module_rows = [
    ["进行中", "黄帝尧", "怪物 AI、波次与交互系统", "本周继续负责怪物更新。围绕第三周“完善敌人 AI、波次刷怪、敌人追击/攻击、受击死亡”的目标，推进敌人配置修复、新敌人视觉资源接入、灯笼修复交互修复和灯光可见性增强。当前已有 ModengEnemy、ModengFastEnemy、ModengRangedEnemy、ModengExploderEnemy、ModengEnemySpawner 等 C++ 基础，后续重点是统一波次节奏、敌人攻击灯笼逻辑和第一关 Alpha 测试口径。"],
    ["进行中", "邢增浩", "主角移动、跳跃与连招攻击", "本周继续负责主角更新。主角侧新增/接入 Tomoe 跳跃下落运动、武士刀攻击检测窗口、地面连招和攻击动画通知流程，包括 Begin/EndAttackTrace、Open/CloseComboWindow、CommitCombo、FinishGroundAttack 等。该部分已经从第二周普通攻击基础推进到更完整的战斗表现，后续需要继续联调玩家血量、受击反馈、灯火弹或击退效果。"],
    ["进行中", "黄顺金", "第一关场景与资源整理", "在第二周第一关原型基础上继续整理灯笼、敌人出生点、玩家出生点和街巷场景资源，配合主角和怪物更新进行关卡摆放与展示效果调整。当前重点是让第一关灰盒向 Alpha 展示场景过渡，并减少资源引用和关卡外部 Actor 冲突。"],
    ["进行中", "张乐钊", "UI、音效与测试支持", "第三周计划要求补齐基础 UI 并接入第一批音效。当前已有 ModengHUD、ModengEnemyHealthWidget、ModengResultWidget 等基础，后续需要把玩家血量、灯笼耐久、当前波次、敌人数量、胜负结果等信息整理成更正式的 UI，并同步记录组内测试问题。"],
    ["进行中", "李远涵", "技术展示与后处理准备", "继续为后续自定义后处理效果做准备。本周灯笼交互和灯光可见性更新为后处理联动提供了更明确的状态来源，后续可基于灯笼耐久/亮度驱动画面冷暖、饱和度和暗角变化。第三周阶段主要配合核心玩法 Alpha 稳定。"],
]
for row in module_rows:
    modules.append(row)
style_range(modules, "A1:D1", title_fill, Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=11), Alignment(horizontal="center", vertical="center", wrap_text=True))
style_range(modules, "A2:D6")
for r in range(2, 7):
    modules[f"A{r}"].fill = progress_fill
    modules.row_dimensions[r].height = 72

# 本周小结
set_widths(summary, [12, 110])
merge(summary, "A1:B1", "本周小结", title_fill, Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=14), Alignment(vertical="center"))
summary.append(["序号", "小组总结"])
summary_rows = [
    ["1", "本周是项目计划中的第三周，开发重点从第二周的最小玩法闭环推进到第一关 Alpha 雏形。团队继续围绕玩家、敌人、灯笼、波次、UI 和表现反馈进行联调，目标是让第一关不仅能跑通，而且能更稳定、更清楚地展示玩法。"],
    ["2", "分工上，本周仍然保持邢增浩负责主角更新、黄帝尧负责怪物更新。邢增浩主要推进 Tomoe 主角运动和战斗表现，包括跳跃下落、武士刀攻击窗口、地面连招和动画通知流程；黄帝尧主要推进怪物相关更新，包括敌人配置、敌人视觉资源、敌人生成/战斗基础，以及灯笼交互与灯光可见性支撑。"],
    ["3", "与第三周验收标准相比，项目已经在敌人资源、怪物配置、主角攻击表现、动画接入和灯笼交互反馈上取得进展；但玩家血量、灯火弹或击退、基础 UI 美术化、第一批音效和组内测试记录仍需继续补齐。下一步应集中完成第一关 Alpha 演示版本，并把风险点记录到测试表中。"],
]
for row in summary_rows:
    summary.append(row)
style_range(summary, "A1:B5")
style_range(summary, "A2:B2", section_fill, header_font, Alignment(horizontal="center", vertical="center"))
for r in range(3, 6):
    summary.row_dimensions[r].height = 70

# 个人小结
set_widths(personal, [18, 110])
personal.append(["组员", "小结"])
personal_rows = [
    ["黄帝尧", "本周继续负责怪物更新，围绕第三周目标推进敌人配置、敌人视觉资源、敌人生成与战斗基础，并同步处理灯笼修复交互、灯光可见性等与怪物攻击目标相关的支撑内容。后续将继续完善波次刷怪、敌人追击/攻击、受击死亡和胜负判定联调。"],
    ["邢增浩", "本周继续负责主角更新，重点推进 Tomoe 主角的跳跃下落运动、武士刀攻击窗口、地面连招和动画通知流程，使主角战斗表现从普通攻击基础进入更完整的连招和动画阶段。后续需要继续接入玩家血量、受击反馈、灯火弹或击退表现。"],
    ["黄顺金", "本周继续配合第一关场景与资源整理，围绕灯笼、敌人出生点、玩家出生点和街巷展示效果进行调整，为第一关 Alpha 演示提供更稳定的关卡基础。"],
    ["张乐钊", "本周继续关注 UI、音效和测试支持。当前已有 HUD、敌人血条和结果界面基础，后续需要把第三周要求的玩家血量、灯笼耐久、波次信息和第一批音效整理进可演示版本，并形成组内测试记录。"],
    ["李远涵", "本周继续准备技术展示方向，关注灯笼状态、亮度和颜色反馈与后处理效果之间的衔接。后续可在第一关 Alpha 稳定后，将灯笼状态映射到画面冷暖、饱和度和暗角变化。"],
]
for row in personal_rows:
    personal.append(row)
style_range(personal, "A1:B6")
style_range(personal, "A1:B1", section_fill, header_font, Alignment(horizontal="center", vertical="center"))
for r in range(2, 7):
    personal.row_dimensions[r].height = 74

# 讨论纪要与截图
set_widths(meeting, [12, 110])
merge(meeting, "A1:B1", "小组讨论纪要和必要截图", title_fill, Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=14), Alignment(vertical="center"))
meeting.append(["序号", "讨论纪要"])
meeting_rows = [
    ["1", "本周讨论确认当前已经进入第三周，应以第一关 Alpha 为目标，不再只描述第二周最小闭环。第三周重点包括敌人 AI、波次刷怪、玩家血量、投射物或击退、基础 UI、音效和测试记录。"],
    ["2", "团队确认本周主要分工：邢增浩继续负责主角更新，重点处理 Tomoe 主角运动、攻击窗口和地面连招；黄帝尧继续负责怪物更新，重点处理敌人配置、视觉资源、生成与战斗表现。其他成员围绕场景、UI、测试和技术展示进行支撑。"],
    ["3", "本周 GitHub 更新主要包括 Tomoe jump/fall locomotion、katana trace attack window、montage-driven ground combo、melee enemy loadouts 修复、新敌人视觉和灯笼资源、灯笼修复交互修复与灯光可见性增强。"],
    ["4", "后续协作仍需注意 Unreal 资源冲突，尤其是 BP_SideScrollingCharacter、ABP_Tomoe_SideScroller、BP_ModengEnemy、BP_ModengEnemySpawner、L_Level01_Street 和 Content/__ExternalActors__ 下的关卡外部 Actor 文件，修改前最好先在群里同步。"],
]
for row in meeting_rows:
    meeting.append(row)
meeting.append(["", ""])
meeting.append(["必要截图清单", ""])
screenshots = [
    ["1", "Tomoe 主角跳跃/下落或移动动画运行截图。"],
    ["2", "主角武士刀攻击窗口或地面连招演示截图。"],
    ["3", "怪物生成、追击或攻击灯笼截图。"],
    ["4", "新敌人视觉资源或怪物配置更新截图。"],
    ["5", "灯笼修复交互与灯光可见性增强截图。"],
    ["6", "基础 HUD / 敌人血条 / 胜负界面截图。"],
]
for row in screenshots:
    meeting.append(row)
style_range(meeting, "A1:B14")
style_range(meeting, "A2:B2", section_fill, header_font, Alignment(horizontal="center", vertical="center"))
style_range(meeting, "A8:B8", section_fill, header_font)
for r in range(3, 7):
    meeting.row_dimensions[r].height = 58
for r in range(9, 15):
    meeting.row_dimensions[r].height = 34

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = copy(cell.font)
            if cell.value is not None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

wb.save(output)
wb.save(workspace_output)
print(output)
